import openpyxl
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
try:
    wb = openpyxl.load_workbook('G:/05_KIEM DINH/data/Kiểm định Master Test.xlsx')
    print('Sheets:', wb.sheetnames)
    ws = wb['Master thiết bị']
    print('Max row:', ws.max_row)
    for row in range(1, 6):
        vals = [ws.cell(row, col).value for col in range(1, 8)]
        print(vals)
except Exception as e:
    print('Error:', e)