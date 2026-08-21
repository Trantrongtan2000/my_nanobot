import openpyxl
import sys
sys.stdout = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1)

wb = openpyxl.load_workbook('G:/05_KIEM DINH/data/Kiểm định Master Test.xlsx')
# Check all sheet names
print('Sheets:', wb.sheetnames)

# Check Master thiết bị sheet
if 'Master thiết bị' in wb.sheetnames:
    ws = wb['Master thiết bị']
    print('\nMaster thiết bị sheet:')
    print('Rows:', ws.max_row, 'Cols:', ws.max_column)
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
    print('Headers:', headers)
    for r in range(2, min(10, ws.max_row+1)):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
        print(f'Row {r}:', row)