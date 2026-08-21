import openpyxl
import sys
sys.stdout = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1)

wb = openpyxl.load_workbook('G:/05_KIEM DINH/data/Kiểm định Master Test.xlsx')
ws = wb.active
print('Sheet:', ws.title)
print('Rows:', ws.max_row, 'Cols:', ws.max_column)
headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
print('Headers:', headers)
# Print some actual data rows (skip the formula rows)
for r in range(10, 20):
    row = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
    print(f'Row {r}:', row)