import openpyxl
import sys
sys.stdout = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1)

wb = openpyxl.load_workbook('G:/05_KIEM DINH/data/Kiểm định Master Test.xlsx')
ws = wb['Master thiết bị']

print("=== Devices with SN starting with '(21)' ===")
count = 0
for row in range(2, ws.max_row + 1):
    sn = ws.cell(row, 1).value
    if sn and str(sn).startswith('(21)'):
        count += 1
        device_name = ws.cell(row, 4).value
        model = ws.cell(row, 5).value
        dept = ws.cell(row, 2).value
        room = ws.cell(row, 3).value
        print(f"Row {row}: SN={sn}, Dept={dept}, Room={room}, Device={device_name}, Model={model}")
        if count >= 10:  # Show first 10
            print("... (showing first 10)")
            break

print(f"\nTotal devices with SN starting '(21)': {count}")

print("\n=== Device with SN '0000507245-004' ===")
for row in range(2, ws.max_row + 1):
    sn = ws.cell(row, 1).value
    if sn and str(sn) == '0000507245-004':
        device_name = ws.cell(row, 4).value
        model = ws.cell(row, 5).value
        dept = ws.cell(row, 2).value
        room = ws.cell(row, 3).value
        print(f"Row {row}: SN={sn}, Dept={dept}, Room={room}, Device={device_name}, Model={model}")
        break