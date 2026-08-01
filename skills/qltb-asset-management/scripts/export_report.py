#!/usr/bin/env python3
"""Export QLTB asset report to Excel/CSV/JSON."""

import argparse
import sqlite3
import json
from datetime import datetime

DB_PATH = "/home/tan/.nanobot/workspace/data/qltb_assets.db"

def export_excel(output_path):
    """Export to Excel using openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl not installed. Install: pip install openpyxl")
        return False
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM assets ORDER BY department, asset_tag")
    rows = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "QLTB Assets"
    
    # Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    for col, name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    
    wb.save(output_path)
    print(f"Excel exported: {output_path} ({len(rows)} assets)")
    return True

def export_csv(output_path):
    """Export to CSV."""
    import csv
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM assets ORDER BY department, asset_tag")
    rows = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]
    conn.close()
    
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(columns)
        writer.writerows(rows)
    
    print(f"CSV exported: {output_path} ({len(rows)} assets)")
    return True

def export_json(output_path):
    """Export to JSON."""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM assets ORDER BY department, asset_tag")
    rows = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]
    conn.close()
    
    data = [dict(zip(columns, row)) for row in rows]
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    print(f"JSON exported: {output_path} ({len(rows)} assets)")
    return True

def main():
    parser = argparse.ArgumentParser(description='Export QLTB asset report')
    parser.add_argument('--format', required=True, choices=['excel', 'csv', 'json'],
                        help='Output format')
    parser.add_argument('--output', required=True, help='Output file path')
    
    args = parser.parse_args()
    
    if args.format == 'excel':
        export_excel(args.output)
    elif args.format == 'csv':
        export_csv(args.output)
    elif args.format == 'json':
        export_json(args.output)

if __name__ == '__main__':
    main()
