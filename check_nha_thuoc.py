import openpyxl
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('G:/05_KIEM DINH/data/Kiểm định Master Test.xlsx')
ws = wb['Master thiết bị']

print("Columns:", [ws.cell(1, col).value for col in range(1, 8)])
print("\nNhà thuốc devices:")
for row in range(2, ws.max_row + 1):
    dept = ws.cell(row, 2).value
    if dept and 'NHÀ THUỐC' in str(dept).upper():
        sn = ws.cell(row, 1).value
        room = ws.cell(row, 3).value
        name = ws.cell(row, 4).value
        model = ws.cell(row, 5).value
        print(f"{sn} | {room} | {name} | {model}")